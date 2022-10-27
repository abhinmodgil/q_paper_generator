import os
import random
from openpyxl import load_workbook
from docx import Document
from docxtpl import DocxTemplate
main_path=r"D:\Question_paper_generate"
template_path= os.path.join(main_path, 'paper_template.docx')
print(template_path)
workbook_path=os.path.join(main_path,'c_question_bank.xlsx')
if __name__ == '__main__':
    workbook= load_workbook(workbook_path)
    template=DocxTemplate(template_path)
    worksheet=workbook["Input"]
    total_records=worksheet.max_row
    sr_no=[]
    section_list = {"A": None, "B": None, "C": None, "D": None}
    sections=["A","B","C","D"]
    for sec in sections:
        questions_2 = []
        sr_no_2 = []
        questions_5 = []
        sr_no_5 = []
        questions_10 = []
        sr_no_10 = []
        questions_20 = []
        sr_no_20 = []
        marks2_list = None
        marks5_list = None
        marks10_list = None
        marks20_list = None
        marks_list = {2: None, 5: None, 10: None, 20: None}
        #print(sec)
        for row in range(2, total_records + 1):
            if worksheet.cell(column=3, row=row).value == sec and worksheet.cell(column=4, row=row).value == 2:
                questions_2.append(worksheet.cell(column=2, row=row).value)
                sr_no_2.append(worksheet.cell(column=1, row=row).value)
                marks2_list = dict(zip(sr_no_2, questions_2))
            if worksheet.cell(column=3, row=row).value == sec and worksheet.cell(column=4, row=row).value == 5:
                questions_5.append(worksheet.cell(column=2, row=row).value)
                sr_no_5.append(worksheet.cell(column=1, row=row).value)
                marks5_list = dict(zip(sr_no_5, questions_5))
            if worksheet.cell(column=3, row=row).value == sec and worksheet.cell(column=4, row=row).value == 10:
                questions_10.append(worksheet.cell(column=2, row=row).value)
                sr_no_10.append(worksheet.cell(column=1, row=row).value)
                marks10_list = dict(zip(sr_no_10, questions_10))
            if worksheet.cell(column=3, row=row).value == sec and worksheet.cell(column=4, row=row).value == 20:
                questions_20.append(worksheet.cell(column=2, row=row).value)
                sr_no_20.append(worksheet.cell(column=1, row=row).value)
                marks20_list = dict(zip(sr_no_20, questions_20))
        marks_list[2] = marks2_list
        marks_list[5] = marks5_list
        marks_list[10] = marks10_list
        marks_list[20] = marks20_list
        #sr_no = sr_no_2 + sr_no_5 + sr_no_10 + sr_no_20
        section_list[sec] = marks_list
    print(section_list)
    marks=[2,5,10,20]
    for sec in sections:
        for mark in marks:
            print(f"\n{mark} Marks questions for section-{sec} are:")
            guess_2_list = set()
            if mark==2:
                for i in range(1, 3):
                    #print(list((section_list[sec][mark].keys())))
                    guessed_number = random.choice(list(section_list[sec][mark].keys()))
                    guess_2_list.add(guessed_number)
                #print(guess_2_list)
                q_section_A = set()
                for q in guess_2_list:
                    if q in section_list[sec][mark].keys():
                        q_section_A.add(section_list[sec][mark][q])
                        print(section_list[sec][mark][q], end="")
                    print("2 Marks")
            else:
                for i in range(1, 5):
                    guessed_number = random.choice(list(section_list[sec][mark].keys()))
                    guess_2_list.add(guessed_number)
                    #print(guess_2_list)
                    q_section_A = set()
                for q in guess_2_list:
                    if q in section_list[sec][mark].keys():
                        q_section_A.add(section_list[sec][mark][q])
                        print(section_list[sec][mark][q], end="")
                    print(f"{mark} Marks")

